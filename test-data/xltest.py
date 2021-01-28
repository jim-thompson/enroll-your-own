'''
Created on Jan 19, 2021

@author: jct
'''

import re
from openpyxl import load_workbook

reprog_first_name = re.compile("first", flags=re.IGNORECASE)
reprog_last_name =  re.compile("last", flags=re.IGNORECASE)
reprog_email = re.compile("e(-)?mail", flags=re.IGNORECASE)
reprog_organization =   re.compile("(company)|(organization)", flags=re.IGNORECASE)

def looks_like(value, prog):
    if type(value) != type('string'):
        return False
    
    match = prog.search(value)
    
    if match is not None:
        return True
    
    return False    

def looks_like_first_name(value):
    return looks_like(value, reprog_first_name)

def looks_like_last_name(value):
    return looks_like(value, reprog_last_name)

def looks_like_email(value):
    return looks_like(value, reprog_email)

def looks_like_organization(value):
    return looks_like(value, reprog_organization)

def looks_like_header(row):
    found_first = False
    found_last = False
    found_email = False
    
    for cell in row:
        if cell is None:
            continue
        
        if looks_like_first_name(cell):
            found_first = True
        elif looks_like_last_name(cell):
            found_last = True
        elif looks_like_email(cell):
            found_email = True
            
        if found_first and found_last and found_email:
            return True
        
    return False

def identify_columns(row):
    columns = {}
    column_num = 0
    email_count = 0
    
    # Initialize the key for email2 so a lookup failure won't
    # trigger an exception. Failure to lookup any other key
    # SHOULD trigger an exception. 
    columns["email2"] = 0
    
    for cell in row:
        if looks_like_first_name(cell):
            columns["first-name"] = column_num
        elif looks_like_last_name(cell):
            columns["last-name"] = column_num
        elif looks_like_email(cell):
            # At this point it becomes more difficult
            if email_count == 0:
                columns["email1"] = column_num
            else:
                columns["email2"] = column_num
            email_count += 1
        elif looks_like_organization(cell):
            columns["organization"] = column_num
            
        column_num += 1

    return columns

def dump_headers(columns):
    for key in columns.keys():
        value = columns[key]
        print(key, value)

def identify_header(wb):
    ws = wb.active
    row_number = 1
    for row in ws.values:
        if looks_like_header(row):
            print("Found a header row:")
            for cell in row:
                print(cell, ' ', end='')
            print()
            
            columns = identify_columns(row)
            
            return columns, row_number
        
        row_number += 1
    
    print("DIDN'T FIND A HEADER ROW :(")
    return None
    
reprog_ocse = re.compile('ocse', flags=re.IGNORECASE)
reprog_nawcwd = re.compile('nawcwd', flags=re.IGNORECASE)
reprog_leidos = re.compile('enrollment request #[0-9]+', flags=re.IGNORECASE)
    
def guess_org(name):
    if reprog_ocse.search(name) is not None:
        return "US Army Office of the Chief Systems Engineer"
    elif reprog_nawcwd.search(name) is not None:
        return "NAWCWD"
    elif reprog_leidos.search(name) is not None:
        return "Leidos" 
    else:
        return None

def load_spreadsheet_contents(ws, columns, first_row, override_org):
    
    content = []
    
    first_name_col = columns['first-name']
    last_name_col = columns['last-name']
    email1_col = columns['email1']
    email2_col = columns['email2']
    
    primary_email = None
    secondary_email = None
    
    if override_org is None:
        organization_col = columns['organization']

    for row in ws.iter_rows(min_row = first_row + 1):
        #row = ws[first_row + 1]
#         cell = row[3]
#         it = cell.value
        
        first_name = row[first_name_col].value
        last_name = row[last_name_col].value
        email1 = row[email1_col].value
        
        if email2_col == 0:
            email2 = None
        else:
            email2 = row[email2_col].value
            
        if email1 is not None:
            primary_email = email1.strip()
            if email2 is not None:
                secondary_email = email2.strip()
        elif email2 is not None:
            primary_email = email2.strip()
                        
        # At this point, primary_email must have a value,
        # and secondary_email may have a value.
        
        if override_org is None:
            organization = row[organization_col].value
        else:
            organization = override_org
        
        # These four fields must now have values (secondary_email can be None):
        if first_name is None or \
           last_name is None or \
           primary_email is None or \
           organization is None:
                continue
            
        val_dict = {}
        val_dict["first-name"] = first_name
        val_dict["last-name"] = last_name
        val_dict['organization'] = organization
        val_dict['primary-email'] = primary_email
        val_dict['secondary-email'] = secondary_email
        
        content.append(val_dict)
        
    return content
        

def load_spreadsheet(name):
    try:
        print()
        print('===========================================================================')
        print("Opening <%s>" % name)
        wb = load_workbook(name)
        ws = wb.active
        
        print('---------------------------------------------------------------------------')
        
        org = guess_org(name)
        print("Guessed org = <%s>" % org)
        
        columns, first_row = identify_header(wb)
        
        if columns is not None:
            print("First row = %d" % first_row)
            dump_headers(columns)
            content = load_spreadsheet_contents(ws, columns, first_row, org)
            
            for row in content:
                first_name = row["first-name"]
                last_name = row['last-name']
                organization = row['organization']
                primary_email = row['primary-email']
                secondary_email = row['secondary-email']
            
                print("---> <%s> <%s> <%s> <%s> <%s>" 
                      % (first_name, last_name, organization, primary_email, secondary_email))
        
            return content
            
    except:
        print("*** error opening spreadsheet.")
    
    return None

if __name__ == '__main__':
    print("*** start ***")
    load_spreadsheet("Add Participants SysML Online Course_Booz Allen_PO D377577_Invoice0001634 (20210119).xlsx")
    load_spreadsheet('Add Participants, SysML Online Course_NAWCAD_Inv 1013_20210119.csv.xlsx')
    load_spreadsheet('Add Participants, SysML Online Course_SAIC_1622_12.28.2020.xlsx')
    load_spreadsheet('Add-MBSE-Subscription-L3Harris-Invoice1983-{01-13-21}.xlsx')
    load_spreadsheet('Add-SysML-Lifetime-OCSE-round2_11Dec2020.xlsx')
    load_spreadsheet('Add-SysML-Subscription-NAWCWD-Invoice1878-13JAN2021.xlsx')
    load_spreadsheet('Enrollment Request #57 - 1 sent 930am 21 December 2020.xlsx')
    load_spreadsheet('Add-SysML-Lifetime-LockheedMartin-Invoice1720-{1-12-21}.xlsx')
    